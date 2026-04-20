import os
import shutil
import math
from typing import List, Optional
from pydantic import BaseModel
from fastapi import FastAPI, Depends, HTTPException, Query, File, UploadFile, Form, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session, joinedload
from sqlalchemy.sql.expression import func
import io
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import models
import schemas
from database import SessionLocal, engine
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime

# --- CẤU HÌNH HỆ THỐNG ---
models.Base.metadata.create_all(bind=engine)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "static", "images")
os.makedirs(UPLOAD_DIR, exist_ok=True)
# --- CẤU HÌNH EMAIL (dùng Gmail App Password) ---
EMAIL_HOST = 'sandbox.smtp.mailtrap.io'
EMAIL_PORT = 2525           
EMAIL_USER = 'f6a814419913ed'
EMAIL_PASS = 'add94cf5de6edd'
EMAIL_NAME = "BookStore Support"
app = FastAPI(
    title="Hệ thống Quản lý Nhà sách - Team Bảo",
    description="API quản lý sách tích hợp phân quyền và bảo mật người dùng",
    version="1.8.0"
)

# --- CẤU HÌNH CORS (PHẢI ĐẶT TRƯỚC CÁC ROUTE) ---
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Cấu hình Static Files
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")

# --- DEPENDENCIES & HELPERS ---
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def check_admin_role(user_role: Optional[str]):
    if user_role != "admin":
        raise HTTPException(status_code=403, detail="Bạn không có quyền Quản trị viên!")

class CategoryCreate(BaseModel):
    name: str
class ContactCreate(BaseModel):
    name: str
    email: str
    subject: str
    message: str

class ReplyCreate(BaseModel):
    reply_message: str
# --- 1. HỆ THỐNG & TÀI KHOẢN ---

@app.get("/", tags=["Hệ thống"])
def read_root():
    return {"message": "Backend Nhà sách Team Bảo đang hoạt động!", "status": "Online"}

@app.post("/login", tags=["Tài khoản"])
def login(user_data: schemas.UserLogin, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.username == user_data.username).first()
    if not user or user.password != user_data.password:
        raise HTTPException(status_code=400, detail="Sai tài khoản hoặc mật khẩu")
    return {
        "id": user.id,
        "username": user.username,
        "role": user.role,
        "name": user.username
    }

@app.post("/users/", response_model=schemas.UserResponse, tags=["Tài khoản"])
def create_user(user: schemas.UserCreate, db: Session = Depends(get_db)):
    db_user = db.query(models.User).filter(models.User.username == user.username).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Tên đăng nhập đã tồn tại!")
    new_user = models.User(**user.model_dump())
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user

# --- 2. QUẢN LÝ DANH MỤC ---

@app.get("/categories/", response_model=List[schemas.CategoryResponse], tags=["Danh mục"])
def list_categories(db: Session = Depends(get_db)):
    return db.query(models.Category).all()

@app.post("/categories/", tags=["Danh mục"])
def create_category(
    cat: CategoryCreate, 
    db: Session = Depends(get_db),
    x_user_role: Optional[str] = Header(None)
):
    check_admin_role(x_user_role)
    new_cat = models.Category(name=cat.name)
    db.add(new_cat)
    db.commit()
    db.refresh(new_cat)
    return new_cat

@app.put("/categories/{cat_id}/", tags=["Danh mục"])
def update_category(
    cat_id: int, 
    cat: CategoryCreate, 
    x_user_role: Optional[str] = Header(None), 
    db: Session = Depends(get_db)
):
    check_admin_role(x_user_role)
    db_cat = db.query(models.Category).filter(models.Category.id == cat_id).first()
    if not db_cat:
        raise HTTPException(status_code=404, detail="Không tìm thấy danh mục")
    db_cat.name = cat.name
    db.commit()
    return db_cat

@app.delete("/categories/{cat_id}/", tags=["Danh mục"])
def delete_category(
    cat_id: int, 
    x_user_role: Optional[str] = Header(None), 
    db: Session = Depends(get_db)
):
    check_admin_role(x_user_role)
    db_cat = db.query(models.Category).filter(models.Category.id == cat_id).first()
    if not db_cat:
        raise HTTPException(status_code=404, detail="Không tìm thấy danh mục")
    
    # Ràng buộc: Không xóa danh mục nếu có sách
    if db.query(models.Book).filter(models.Book.category_id == cat_id).count() > 0:
        raise HTTPException(status_code=400, detail="Danh mục có sách, không thể xóa!")
        
    db.delete(db_cat)
    db.commit()
    return {"message": "Xóa danh mục thành công"}

# --- 3. QUẢN LÝ SÁCH ---

@app.get("/books/", tags=["Sách"])
def get_books(
    page: int = 1, limit: int = 20, 
    search: Optional[str] = None, 
    category: Optional[str] = None, 
    db: Session = Depends(get_db)
):
    query = db.query(models.Book)
    if search:
        query = query.filter(models.Book.title.icontains(search))
    if category and category.strip() != "":
        query = query.join(models.Category).filter(models.Category.name == category)
        
    total_count = query.count()
    books = query.offset((page - 1) * limit).limit(limit).all()
    return {
        "books": books,
        "total_pages": math.ceil(total_count / limit) if limit > 0 else 1,
        "total_items": total_count
    }

@app.post("/books/", tags=["Quản trị - Sách"])
async def create_book(
    title: str = Form(...), author: str = Form(...), price: int = Form(...), 
    stock: int = Form(...), description: str = Form(""), category_id: int = Form(...),
    image: UploadFile = File(...), db: Session = Depends(get_db), 
    x_user_role: Optional[str] = Header(None)
):
    check_admin_role(x_user_role)
    file_path = os.path.join(UPLOAD_DIR, image.filename)
    with open(file_path, "wb+") as buffer:
        shutil.copyfileobj(image.file, buffer)
    
    new_book = models.Book(
        title=title, author=author, price=price, stock=stock,
        description=description, category_id=category_id,
        image_url=f"/static/images/{image.filename}"
    )
    db.add(new_book)
    db.commit()
    db.refresh(new_book)
    return new_book

@app.put("/books/{book_id}/", tags=["Quản trị - Sách"])
async def update_book(
    book_id: int, title: str = Form(...), author: str = Form(...), 
    price: int = Form(...), stock: int = Form(...), description: str = Form(""), 
    category_id: int = Form(...), image: Optional[UploadFile] = File(None), 
    db: Session = Depends(get_db), x_user_role: Optional[str] = Header(None)
):
    check_admin_role(x_user_role)
    db_book = db.query(models.Book).filter(models.Book.id == book_id).first()
    if not db_book:
        raise HTTPException(status_code=404, detail="Không tìm thấy sách")
    
    db_book.title, db_book.author, db_book.price = title, author, price
    db_book.stock, db_book.description, db_book.category_id = stock, description, category_id

    if image:
        file_path = os.path.join(UPLOAD_DIR, image.filename)
        with open(file_path, "wb+") as buffer:
            shutil.copyfileobj(image.file, buffer)
        db_book.image_url = f"/static/images/{image.filename}"

    db.commit()
    return db_book

@app.delete("/books/{book_id}/", tags=["Quản trị - Sách"])
def delete_book(book_id: int, x_user_role: Optional[str] = Header(None), db: Session = Depends(get_db)):
    check_admin_role(x_user_role)
    db_book = db.query(models.Book).filter(models.Book.id == book_id).first()
    if not db_book:
        raise HTTPException(status_code=404, detail="Không tìm thấy sách")
    db.delete(db_book)
    db.commit()
    return {"message": "Xóa sách thành công"}
@app.get("/books/{book_id}", tags=["Sách"])
def get_book_detail(book_id: int, db: Session = Depends(get_db)):
    db_book = db.query(models.Book).filter(models.Book.id == book_id).first()
    if not db_book:
        raise HTTPException(status_code=404, detail="Không tìm thấy cuốn sách này!")
    return db_book
@app.get("/books/random/", tags=["Sách"])
def get_random_books(limit: int = 5, db: Session = Depends(get_db)):
    try:
        books = db.query(models.Book).order_by(func.random()).limit(limit).all()
        return books
    except Exception as e:
        print(f"Lỗi Backend: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# --- XUẤT KHO ra Excel ---
@app.get("/admin/books/export", tags=["Quản trị - Sách"])
def export_books_excel(
    x_user_role: Optional[str] = Header(None),
    db: Session = Depends(get_db)
):
    check_admin_role(x_user_role)
    books = db.query(models.Book).options(joinedload(models.Book.category)).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách sách"

    # Style header
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["ID", "Tên sách", "Tác giả", "Danh mục", "Giá (VNĐ)", "Tồn kho", "Mô tả"]
    col_widths = [6, 40, 25, 20, 15, 12, 50]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 30

    # Dữ liệu
    alt_fill = PatternFill("solid", fgColor="F5F3FF")
    for i, book in enumerate(books, 2):
        row_data = [
            book.id,
            book.title,
            book.author,
            book.category.name if book.category else "",
            book.price,
            book.stock,
            book.description or ""
        ]
        fill = alt_fill if i % 2 == 0 else None
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 7))
            if fill:
                cell.fill = fill
            if col == 5:  # Giá
                cell.number_format = '#,##0'
        ws.row_dimensions[i].height = 20

    # Footer tổng kết
    footer_row = len(books) + 2
    ws.cell(row=footer_row, column=5, value=f"=SUM(E2:E{len(books)+1})").number_format = '#,##0'
    ws.cell(row=footer_row, column=6, value=f"=SUM(F2:F{len(books)+1})")
    label_cell = ws.cell(row=footer_row, column=4, value="TỔNG CỘNG")
    label_cell.font = Font(bold=True, color="4F46E5")
    label_cell.alignment = Alignment(horizontal="right")

    # Freeze header row
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=danh_sach_sach.xlsx"}
    )


# --- NHẬP KHO từ Excel ---
@app.post("/admin/books/import", tags=["Quản trị - Sách"])
async def import_books_excel(
    file: UploadFile = File(...),
    x_user_role: Optional[str] = Header(None),
    db: Session = Depends(get_db)
):
    check_admin_role(x_user_role)

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Chỉ chấp nhận file .xlsx hoặc .xls!")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    results = {"updated": 0, "not_found": [], "errors": []}

    # Bỏ qua hàng header (row 1), đọc từ row 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            book_id = int(row[0])
            new_stock = row[5]  # Cột F = Tồn kho (index 5)

            if new_stock is None:
                continue

            new_stock = int(new_stock)
            if new_stock < 0:
                results["errors"].append(f"ID {book_id}: Tồn kho không được âm")
                continue

            db_book = db.query(models.Book).filter(models.Book.id == book_id).first()
            if not db_book:
                results["not_found"].append(book_id)
                continue

            db_book.stock = new_stock
            results["updated"] += 1

        except (ValueError, TypeError) as e:
            results["errors"].append(f"Lỗi dòng {row}: {str(e)}")

    db.commit()
    return {
        "message": f"Nhập kho hoàn tất! Đã cập nhật {results['updated']} sách.",
        **results
    }
# --- 4. GIỎ HÀNG  ---

@app.post("/cart/", response_model=schemas.CartItemResponse, tags=["Giỏ hàng"])
def add_to_cart(item: schemas.CartItemCreate, db: Session = Depends(get_db)):
    db_item = db.query(models.CartItem).filter(
        models.CartItem.book_id == item.book_id, 
        models.CartItem.user_id == item.user_id
    ).first()
    if db_item:
        db_item.quantity += item.quantity
    else:
        db_item = models.CartItem(**item.model_dump())
        db.add(db_item)
    db.commit()
    db.refresh(db_item)
    return db_item

@app.get("/cart/{user_id}", response_model=List[schemas.CartItemResponse], tags=["Giỏ hàng"])
def get_user_cart(user_id: int, db: Session = Depends(get_db)):
    return db.query(models.CartItem).options(joinedload(models.CartItem.book)).filter(models.CartItem.user_id == user_id).all()

@app.put("/cart/{item_id}", response_model=schemas.CartItemResponse, tags=["Giỏ hàng"])
def update_cart_quantity(item_id: int, update_data: schemas.CartUpdateQuantity, db: Session = Depends(get_db)):
    db_item = db.query(models.CartItem).filter(models.CartItem.id == item_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Không tìm thấy mục này trong giỏ hàng")
    
    # Kiểm tra tồn kho của sách đó
    book = db.query(models.Book).filter(models.Book.id == db_item.book_id).first()
    if update_data.quantity > book.stock:
        raise HTTPException(status_code=400, detail=f"Rất tiếc, kho chỉ còn {book.stock} cuốn!")

    db_item.quantity = update_data.quantity
    db.commit()
    db.refresh(db_item)
    return db_item

@app.delete("/cart/{item_id}", tags=["Giỏ hàng"])
def delete_cart_item(item_id: int, db: Session = Depends(get_db)):
    db_item = db.query(models.CartItem).filter(models.CartItem.id == item_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Mục này không tồn tại")
    
    db.delete(db_item)
    db.commit()
    return {"message": "Đã xóa khỏi giỏ hàng thành công"}
# --- 5. ĐƠN HÀNG (USER & ADMIN) ---

@app.post("/orders/{user_id}", tags=["Đơn hàng"])
def create_order(user_id: int, db: Session = Depends(get_db)):
    cart_items = db.query(models.CartItem).options(joinedload(models.CartItem.book)).filter(models.CartItem.user_id == user_id).all()
    if not cart_items:
        raise HTTPException(status_code=400, detail="Giỏ hàng trống!")

    # ── Kiểm tra tồn kho trước khi tạo đơn ──
    for item in cart_items:
        if item.book.stock < item.quantity:
            raise HTTPException(
                status_code=400,
                detail=f"Sách '{item.book.title}' chỉ còn {item.book.stock} cuốn trong kho!"
            )

    total = sum(item.book.price * item.quantity for item in cart_items)
    new_order = models.Order(user_id=user_id, total_price=total, status="Success")
    db.add(new_order)
    db.flush()

    for item in cart_items:
        db.add(models.OrderItem(
            order_id=new_order.id, book_id=item.book_id,
            quantity=item.quantity, price_at_purchase=item.book.price
        ))
        # ── Trừ tồn kho ──
        item.book.stock -= item.quantity

    db.query(models.CartItem).filter(models.CartItem.user_id == user_id).delete()
    db.commit()
    return {"message": "Thanh toán thành công!", "order_id": new_order.id, "total": total}

@app.get("/orders/{user_id}", tags=["Đơn hàng"])
def get_user_order_history(user_id: int, db: Session = Depends(get_db)):
    return db.query(models.Order).options(
        joinedload(models.Order.items).joinedload(models.OrderItem.book)
    ).filter(models.Order.user_id == user_id).order_by(models.Order.id.desc()).all()

@app.get("/admin/orders/", tags=["Quản trị - Đơn hàng"])
def get_all_orders_admin(x_user_role: Optional[str] = Header(None), db: Session = Depends(get_db)):
    check_admin_role(x_user_role)
    # Thêm joinedload(models.Order.owner) vào đây
    return db.query(models.Order).options(
        joinedload(models.Order.owner), # Lấy thông tin khách hàng
        joinedload(models.Order.items).joinedload(models.OrderItem.book)
    ).order_by(models.Order.id.desc()).all()

# --- 6. QUẢN TRỊ NGƯỜI DÙNG ---

@app.get("/admin/users/", tags=["Quản trị - Người dùng"])
def list_all_users_admin(x_user_role: Optional[str] = Header(None), db: Session = Depends(get_db)):
    check_admin_role(x_user_role)
    return db.query(models.User).all()

@app.post("/admin/users/", response_model=schemas.UserResponse, tags=["Quản trị - Người dùng"])
def admin_create_user(user: schemas.UserCreate, db: Session = Depends(get_db), x_user_role: Optional[str] = Header(None)):
    check_admin_role(x_user_role)
    db_user = db.query(models.User).filter(models.User.username == user.username).first()
    if db_user:
         raise HTTPException(status_code=400, detail="Tên đăng nhập đã tồn tại!")
    new_user = models.User(**user.model_dump())
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user

@app.put("/admin/users/{user_id}", tags=["Quản trị - Người dùng"])
def admin_update_user(user_id: int, user_data: schemas.UserCreate, x_user_role: Optional[str] = Header(None), db: Session = Depends(get_db)):
    check_admin_role(x_user_role)
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="Không tìm thấy người dùng")
    
    # Cập nhật thông tin
    db_user.username = user_data.username
    db_user.password = user_data.password # Lưu ý: Trong thực tế nên hash mật khẩu
    db_user.role = user_data.role
    
    db.commit()
    db.refresh(db_user)
    return db_user

@app.delete("/admin/users/{user_id}", tags=["Quản trị - Người dùng"])
def admin_delete_user(user_id: int, x_user_role: Optional[str] = Header(None), db: Session = Depends(get_db)):
    check_admin_role(x_user_role)
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="Không tìm thấy người dùng")
    
    db.delete(db_user)
    db.commit()
    return {"message": "Xóa người dùng thành công"}
# --- 7. Wishlist ---

@app.get("/wishlist/{user_id}", tags=["Yêu thích"])
def get_wishlist(user_id: int, db: Session = Depends(get_db)):
    items = db.query(models.Wishlist)\
              .options(joinedload(models.Wishlist.book))\
              .filter(models.Wishlist.user_id == user_id)\
              .all()
    return items

@app.post("/wishlist/toggle", tags=["Yêu thích"])
def toggle_wishlist(data: dict, db: Session = Depends(get_db)):
    user_id = data.get("user_id")
    book_id = data.get("book_id")
    item = db.query(models.Wishlist).filter(
        models.Wishlist.user_id == user_id, 
        models.Wishlist.book_id == book_id
    ).first()

    if item:
        db.delete(item)
        db.commit()
        return {"status": False, "message": "Đã xóa khỏi danh sách yêu thích"}
    else:
        new_fav = models.Wishlist(user_id=user_id, book_id=book_id)
        db.add(new_fav)
        db.commit()
        return {"status": True, "message": "Đã thêm vào danh sách yêu thích"}
# --- 8. Review ---
class ReviewCreate(BaseModel):
    book_id: int
    user_id: int
    rating: int
    comment: str

@app.get("/reviews/{book_id}", tags=["Đánh giá"])
def get_book_reviews(book_id: int, db: Session = Depends(get_db)):
    # Lấy danh sách đánh giá của một cuốn sách
    return db.query(models.Review).filter(models.Review.book_id == book_id).all()

@app.post("/reviews/", tags=["Đánh giá"])
def create_review(review: ReviewCreate, db: Session = Depends(get_db)):
    # 1. Kiểm tra xem user đã đánh giá cuốn sách này chưa
    existing_review = db.query(models.Review).filter(
        models.Review.user_id == review.user_id,
        models.Review.book_id == review.book_id
    ).first()
    
    if existing_review:
        raise HTTPException(
            status_code=400, 
            detail="Bạn đã đánh giá cuốn sách này rồi. Mỗi người chỉ được đánh giá một lần!"
        )
    
    # 2. Kiểm tra tính hợp lệ của số sao
    if review.rating < 1 or review.rating > 5:
        raise HTTPException(status_code=400, detail="Số sao phải từ 1 đến 5")

    # 3. Lưu đánh giá mới
    new_review = models.Review(
        user_id=review.user_id,
        book_id=review.book_id,
        rating=review.rating,
        comment=review.comment
    )
    db.add(new_review)
    db.commit()
    db.refresh(new_review)
    return {"message": "Gửi đánh giá thành công!", "review": new_review}

# --- 9. LIÊN HỆ ---

@app.post("/contact/", tags=["Liên hệ"])
def submit_contact(data: ContactCreate, db: Session = Depends(get_db)):
    new_msg = models.Contact(
        name=data.name,
        email=data.email,
        subject=data.subject,
        message=data.message,
        status="pending",
        created_at=datetime.utcnow()
    )
    db.add(new_msg)
    db.commit()
    db.refresh(new_msg)
    return {"message": "Gửi tin nhắn thành công!", "id": new_msg.id}


@app.get("/admin/contacts/", tags=["Quản trị - Liên hệ"])
def get_all_contacts(
    x_user_role: Optional[str] = Header(None),
    db: Session = Depends(get_db)
):
    check_admin_role(x_user_role)
    return db.query(models.Contact).order_by(models.Contact.id.desc()).all()


@app.post("/admin/contacts/{contact_id}/reply", tags=["Quản trị - Liên hệ"])
def reply_contact(
    contact_id: int,
    data: ReplyCreate,
    x_user_role: Optional[str] = Header(None),
    db: Session = Depends(get_db)
):
    check_admin_role(x_user_role)

    contact = db.query(models.Contact).filter(models.Contact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Không tìm thấy tin nhắn!")

    # Gửi email
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"Re: {contact.subject} — BookStore"
        msg["From"]    = f"{EMAIL_NAME} <{EMAIL_USER}>"
        msg["To"]      = contact.email

        html_body = f"""
        <div style="font-family: sans-serif; max-width: 600px; margin: auto; padding: 32px; background: #f9fafb; border-radius: 16px;">
          <div style="background: #4F46E5; color: white; padding: 24px 32px; border-radius: 12px 12px 0 0;">
            <h1 style="margin:0; font-size:24px; font-style:italic;">BookStore</h1>
          </div>
          <div style="background: white; padding: 32px; border-radius: 0 0 12px 12px; border: 1px solid #e5e7eb; border-top: none;">
            <p style="color:#6b7280;">Xin chào <strong>{contact.name}</strong>,</p>
            <p style="color:#6b7280;">Cảm ơn bạn đã liên hệ với chúng tôi về chủ đề: <strong>{contact.subject}</strong>.</p>
            <div style="background:#f3f4f6; border-left: 4px solid #4F46E5; padding: 16px 20px; border-radius: 8px; margin: 20px 0;">
              <p style="margin:0; color:#374151; white-space:pre-wrap;">{data.reply_message}</p>
            </div>
            <hr style="border:none; border-top:1px solid #e5e7eb; margin: 24px 0;">
            <p style="color:#9ca3af; font-size:12px;">Đây là email tự động từ BookStore. Vui lòng không trả lời email này.<br>
            Nếu cần hỗ trợ thêm, hãy liên hệ lại qua trang web của chúng tôi.</p>
          </div>
        </div>
        """
        msg.attach(MIMEText(html_body, "html"))

        with smtplib.SMTP(EMAIL_HOST, EMAIL_PORT) as server:
            server.ehlo()
            server.starttls()
            server.login(EMAIL_USER, EMAIL_PASS)
            server.sendmail(EMAIL_USER, contact.email, msg.as_string())

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi gửi email: {str(e)}")

    # Cập nhật DB
    contact.status = "replied"
    contact.reply_message = data.reply_message
    db.commit()

    return {"message": f"Đã phản hồi tới {contact.email}"}


@app.delete("/admin/contacts/{contact_id}", tags=["Quản trị - Liên hệ"])
def delete_contact(
    contact_id: int,
    x_user_role: Optional[str] = Header(None),
    db: Session = Depends(get_db)
):
    check_admin_role(x_user_role)
    contact = db.query(models.Contact).filter(models.Contact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Không tìm thấy!")
    db.delete(contact)
    db.commit()
    return {"message": "Đã xóa tin nhắn"}